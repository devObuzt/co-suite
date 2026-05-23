from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook

from api.services.product_bulk_parser import (
    DEFAULT_HEBREW_MAPPING,
    detect_column_mapping,
    fill_missing_image_refs_from_zip,
    match_zip_images,
    parse_workbook,
)


def make_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["שם", "תמונה", "סלוגן", "תיאור המוצר", "מחיר לסט שלם + מע\"מ", "תוספת בכל העיצובים", "הערות"])
    ws.append(["שולחן עבודה", "desk 01.jpg", "עובדים נכון", "שולחן למשרד ביתי", "1290", "להוסיף לוגו", "צבע עץ"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def make_xlsx_with_leading_empty_rows() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([])
    ws.append([None, "", None])
    ws.append(["שם", "תמונה", "סלוגן"])
    ws.append(["כסא משרדי", "chair.png", "יושבים נכון"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def make_xlsx_with_title_before_headers() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([None, "אורא אופיס - יבוא ושיווק ריהוט משרדי"])
    ws.append(["שם", "תמונה", "סלוגן", "תיאור המוצר", "מחיר לסט שלם + מע\"מ", "תוספת בכל העיצובים", "הערות"])
    ws.append(["ASIL", None, "סלוגן", "שולחן עבודה מודרני", "990", "הנחה", None])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def make_zip() -> bytes:
    out = BytesIO()
    with ZipFile(out, "w") as zf:
        zf.writestr("products/desk 01.jpg", b"fake-image")
        zf.writestr("products/chair.png", b"fake-png")
        zf.writestr("products/shelf.webp", b"fake-webp")
        zf.writestr("ignore/readme.txt", b"not image")
    return out.getvalue()


def test_detect_column_mapping_supports_hebrew_headers():
    headers = ["שם", "תמונה", "סלוגן", "תיאור המוצר", "מחיר לסט שלם + מע\"מ", "תוספת בכל העיצובים", "הערות"]

    mapping = detect_column_mapping(headers)

    assert mapping["product_name"] == "שם"
    assert mapping["image_ref"] == "תמונה"
    assert mapping["slogan"] == "סלוגן"
    assert mapping["description"] == "תיאור המוצר"
    assert mapping["price"] == "מחיר לסט שלם + מע\"מ"
    assert mapping["global_addition"] == "תוספת בכל העיצובים"
    assert mapping["notes"] == "הערות"
    assert DEFAULT_HEBREW_MAPPING["שם"] == "product_name"


def test_parse_workbook_returns_normalized_products():
    parsed = parse_workbook(make_xlsx())

    assert parsed.headers[0] == "שם"
    assert parsed.mapping["product_name"] == "שם"
    assert len(parsed.rows) == 1
    row = parsed.rows[0]
    assert row["product_name"] == "שולחן עבודה"
    assert row["image_ref"] == "desk 01.jpg"
    assert row["price"] == "1290"
    assert row["raw_row"]["הערות"] == "צבע עץ"


def test_parse_workbook_uses_first_non_empty_row_as_headers():
    parsed = parse_workbook(make_xlsx_with_leading_empty_rows())

    assert parsed.headers == ["שם", "תמונה", "סלוגן"]
    assert parsed.mapping["product_name"] == "שם"
    assert len(parsed.rows) == 1
    row = parsed.rows[0]
    assert row["row_index"] == 4
    assert row["product_name"] == "כסא משרדי"
    assert row["image_ref"] == "chair.png"
    assert row["slogan"] == "יושבים נכון"
    assert row["raw_row"] == {
        "שם": "כסא משרדי",
        "תמונה": "chair.png",
        "סלוגן": "יושבים נכון",
    }


def test_parse_workbook_skips_title_rows_before_real_headers():
    parsed = parse_workbook(make_xlsx_with_title_before_headers())

    assert parsed.headers[0] == "שם"
    assert parsed.mapping["product_name"] == "שם"
    assert len(parsed.rows) == 1
    row = parsed.rows[0]
    assert row["row_index"] == 3
    assert row["product_name"] == "ASIL"
    assert row["description"] == "שולחן עבודה מודרני"


def test_fill_missing_image_refs_from_zip_matches_product_name():
    parsed = parse_workbook(make_xlsx_with_title_before_headers())
    out = BytesIO()
    with ZipFile(out, "w") as zf:
        zf.writestr("ASIL-1 1.20 + 1.40 + 1.60.jpg", b"fake-image")

    fill_missing_image_refs_from_zip(parsed.rows, out.getvalue())

    assert parsed.rows[0]["image_ref"] == "asil-1 1.20 + 1.40 + 1.60.jpg"


def test_match_zip_images_matches_by_basename_and_ignores_non_images():
    matches = match_zip_images(make_zip(), ["desk 01.jpg", "chair.png", "shelf.webp", "readme.txt", "missing.jpg"])

    assert "desk 01.jpg" in matches
    assert matches["desk 01.jpg"].filename == "products/desk 01.jpg"
    assert matches["desk 01.jpg"].content_type == "image/jpeg"
    assert matches["chair.png"].filename == "products/chair.png"
    assert matches["chair.png"].content_type == "image/png"
    assert matches["shelf.webp"].filename == "products/shelf.webp"
    assert matches["shelf.webp"].content_type == "image/webp"
    assert "readme.txt" not in matches
    assert "missing.jpg" not in matches
