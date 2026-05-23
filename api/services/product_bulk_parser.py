from dataclasses import dataclass
from io import BytesIO
from pathlib import PurePosixPath
import re
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook


DEFAULT_HEBREW_MAPPING = {
    "שם": "product_name",
    "תמונה": "image_ref",
    "סלוגן": "slogan",
    "תיאור המוצר": "description",
    "מחיר לסט שלם + מע\"מ": "price",
    "תוספת בכל העיצובים": "global_addition",
    "הערות": "notes",
}

IMAGE_CONTENT_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
}


@dataclass
class ParsedWorkbook:
    headers: list[str]
    mapping: dict[str, str]
    rows: list[dict[str, Any]]


@dataclass
class ZipImage:
    filename: str
    data: bytes
    content_type: str


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_filename(value: str) -> str:
    cleaned = clean_cell(value).replace("\\", "/")
    return PurePosixPath(cleaned).name.strip().lower()


def detect_column_mapping(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for header in headers:
        target = DEFAULT_HEBREW_MAPPING.get(clean_cell(header))
        if target:
            mapping[target] = header
    return mapping


def _looks_like_header_row(row: tuple[Any, ...]) -> bool:
    headers = [clean_cell(cell) for cell in row]
    mapping = detect_column_mapping(headers)
    return "product_name" in mapping and len(mapping) >= 2


def parse_workbook(xlsx_bytes: bytes, mapping: dict[str, str] | None = None) -> ParsedWorkbook:
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header_row_index = None
    header_row = None
    for idx, row in enumerate(rows):
        if mapping:
            is_header = any(clean_cell(cell) for cell in row)
        else:
            is_header = _looks_like_header_row(row)
        if is_header:
            header_row_index = idx
            header_row = row
            break

    if header_row_index is None or header_row is None:
        return ParsedWorkbook(headers=[], mapping={}, rows=[])

    headers = [clean_cell(cell) for cell in header_row]
    active_mapping = mapping or detect_column_mapping(headers)
    header_index = {header: idx for idx, header in enumerate(headers)}
    data_rows: list[dict[str, Any]] = []

    for row_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        raw_row = {
            header: clean_cell(row[idx] if idx < len(row) else "")
            for header, idx in header_index.items()
        }
        if not any(raw_row.values()):
            continue

        normalized: dict[str, Any] = {"row_index": row_index, "raw_row": raw_row}
        for field, header in active_mapping.items():
            normalized[field] = raw_row.get(header, "")
        normalized.setdefault("product_name", "")
        normalized.setdefault("image_ref", "")
        data_rows.append(normalized)

    return ParsedWorkbook(headers=headers, mapping=active_mapping, rows=data_rows)


def _matchable_text(value: str) -> str:
    return re.sub(r"[^a-z0-9א-ת\u0600-\u06ff]+", " ", clean_cell(value).lower()).strip()


def _zip_image_candidates(zip_bytes: bytes) -> list[str]:
    with ZipFile(BytesIO(zip_bytes)) as zf:
        return [
            info.filename
            for info in zf.infolist()
            if not info.is_dir() and PurePosixPath(info.filename).suffix.lower() in IMAGE_CONTENT_TYPES
        ]


def fill_missing_image_refs_from_zip(rows: list[dict[str, Any]], zip_bytes: bytes) -> None:
    """Fill blank image_ref values by matching product names against ZIP image filenames."""
    candidates = _zip_image_candidates(zip_bytes)
    used: set[str] = set()

    for row in rows:
        if clean_cell(row.get("image_ref")):
            continue

        product_name = _matchable_text(row.get("product_name", ""))
        if not product_name:
            continue

        product_tokens = product_name.split()
        match = None
        for filename in candidates:
            if filename in used:
                continue
            stem = _matchable_text(PurePosixPath(filename).stem)
            if product_name in stem or all(token in stem for token in product_tokens):
                match = filename
                break

        if match:
            row["image_ref"] = normalize_filename(match)
            row.setdefault("raw_row", {})["__matched_image_from_product_name"] = match
            used.add(match)


def _image_content_type(image_format: str | None) -> tuple[str, str]:
    normalized = (image_format or "png").lower()
    if normalized in {"jpg", "jpeg"}:
        return "image/jpeg", "jpg"
    if normalized == "webp":
        return "image/webp", "webp"
    return "image/png", "png"


def fill_missing_images_from_workbook(rows: list[dict[str, Any]], xlsx_bytes: bytes) -> None:
    """Attach embedded worksheet images to rows that have no image reference."""
    if not rows:
        return

    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[0]
    images = []
    for image in getattr(ws, "_images", []):
        try:
            row_index = image.anchor._from.row + 1
            data = image._data()
        except Exception:
            continue
        content_type, extension = _image_content_type(getattr(image, "format", None))
        images.append(
            {
                "row_index": row_index,
                "filename": f"embedded-row-{row_index}.{extension}",
                "data": data,
                "content_type": content_type,
                "size": len(data),
            }
        )

    if not images:
        return

    images.sort(key=lambda item: (item["row_index"], -item["size"]))
    sorted_rows = sorted(rows, key=lambda item: int(item.get("row_index") or 0))
    used: set[int] = set()

    for index, row in enumerate(sorted_rows):
        if clean_cell(row.get("image_ref")):
            continue

        row_index = int(row.get("row_index") or 0)
        next_row_index = (
            int(sorted_rows[index + 1].get("row_index") or 0)
            if index + 1 < len(sorted_rows)
            else 10**9
        )
        candidates = [
            (image_index, image)
            for image_index, image in enumerate(images)
            if image_index not in used and row_index <= image["row_index"] < next_row_index
        ]
        if not candidates:
            continue

        _, selected = max(candidates, key=lambda item: (item[1]["row_index"] == row_index, item[1]["size"]))
        image_index = images.index(selected)
        used.add(image_index)
        row["image_ref"] = selected["filename"]
        row["__embedded_image"] = ZipImage(
            filename=selected["filename"],
            data=selected["data"],
            content_type=selected["content_type"],
        )


def match_zip_images(zip_bytes: bytes, image_refs: list[str]) -> dict[str, ZipImage]:
    wanted = {normalize_filename(ref): ref for ref in image_refs if clean_cell(ref)}
    matches: dict[str, ZipImage] = {}

    with ZipFile(BytesIO(zip_bytes)) as zf:
        for info in zf.infolist():
            suffix = PurePosixPath(info.filename).suffix.lower()
            if info.is_dir() or suffix not in IMAGE_CONTENT_TYPES:
                continue

            original_ref = wanted.get(normalize_filename(info.filename))
            if original_ref and original_ref not in matches:
                matches[original_ref] = ZipImage(
                    filename=info.filename,
                    data=zf.read(info),
                    content_type=IMAGE_CONTENT_TYPES[suffix],
                )

    return matches
